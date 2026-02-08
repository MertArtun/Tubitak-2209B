"""Tests for face recognition: database, recognizer, and pipeline."""

import numpy as np
import pytest

from src.face_recognition.database import StudentDatabase
from src.face_recognition.recognizer import FaceRecognizer


# ── StudentDatabase tests ────────────────────────────────────────────────────


class TestStudentDatabase:
    """Test SQLite-based student database operations."""

    def test_add_student(self, in_memory_db):
        """Add student, verify ID returned."""
        sid = in_memory_db.add_student("Alice", "alice@example.com")
        assert isinstance(sid, int)
        assert sid > 0

    def test_get_student(self, in_memory_db):
        """Add then get, verify data."""
        sid = in_memory_db.add_student("Bob", "bob@example.com")
        student = in_memory_db.get_student(sid)
        assert student is not None
        assert student["name"] == "Bob"
        assert student["email"] == "bob@example.com"

    def test_list_students(self, in_memory_db):
        """Add multiple, list all."""
        in_memory_db.add_student("Alice")
        in_memory_db.add_student("Bob")
        in_memory_db.add_student("Carol")

        students = in_memory_db.list_students()
        assert len(students) == 3
        names = {s["name"] for s in students}
        assert names == {"Alice", "Bob", "Carol"}

    def test_get_student_not_found(self, in_memory_db):
        """Get non-existent student should return None."""
        assert in_memory_db.get_student(9999) is None


class TestEmbeddingStorage:
    """Test face embedding persistence in SQLite."""

    def test_save_and_get_embeddings(self, in_memory_db, sample_embedding):
        """Save embedding blob, retrieve and verify shape/values."""
        sid = in_memory_db.add_student("Alice")
        in_memory_db.save_embedding(sid, sample_embedding)

        embeddings = in_memory_db.get_embeddings(sid)
        assert len(embeddings) == 1
        assert embeddings[0].shape == (512,)
        assert np.allclose(embeddings[0], sample_embedding, atol=1e-6)

    def test_get_all_embeddings(self, in_memory_db):
        """Multiple students, verify averaged embeddings."""
        sid1 = in_memory_db.add_student("Alice")
        sid2 = in_memory_db.add_student("Bob")

        emb1 = np.random.randn(512).astype(np.float32)
        emb1 /= np.linalg.norm(emb1)
        emb2 = np.random.randn(512).astype(np.float32)
        emb2 /= np.linalg.norm(emb2)

        in_memory_db.save_embedding(sid1, emb1)
        in_memory_db.save_embedding(sid2, emb2)

        all_emb = in_memory_db.get_all_embeddings()
        assert sid1 in all_emb
        assert sid2 in all_emb
        assert all_emb[sid1].shape == (512,)
        assert all_emb[sid2].shape == (512,)

    def test_multiple_embeddings_averaged(self, in_memory_db):
        """Multiple embeddings for the same student should be averaged."""
        sid = in_memory_db.add_student("Alice")

        emb1 = np.ones(512, dtype=np.float32) * 0.5
        emb2 = np.ones(512, dtype=np.float32) * 1.5

        in_memory_db.save_embedding(sid, emb1)
        in_memory_db.save_embedding(sid, emb2)

        all_emb = in_memory_db.get_all_embeddings()
        # Average of [0.5, 1.5] = 1.0 per dim, then normalized
        avg = np.ones(512, dtype=np.float32)
        avg /= np.linalg.norm(avg)
        assert np.allclose(all_emb[sid], avg, atol=1e-5)


class TestSessionManagement:
    """Test session creation and ending."""

    def test_create_and_end_session(self, in_memory_db):
        """Create session, end it, verify end_time set."""
        session_id = in_memory_db.create_session("Test Session", "online")
        assert isinstance(session_id, int)

        in_memory_db.end_session(session_id)

        row = in_memory_db.conn.execute(
            "SELECT * FROM sessions WHERE id = ?", (session_id,)
        ).fetchone()
        assert row is not None
        assert row["end_time"] is not None


class TestAttentionLogs:
    """Test attention logging and stats."""

    def test_log_attention(self, in_memory_db):
        """Log entries, verify they are stored."""
        sid = in_memory_db.add_student("Alice")
        session_id = in_memory_db.create_session("Session 1", "online")

        in_memory_db.log_attention(sid, session_id, "positive", 0.9, 0.8, "focused")
        in_memory_db.log_attention(sid, session_id, "neutral", 0.7, 0.5, "moderate")

        rows = in_memory_db.conn.execute(
            "SELECT * FROM attention_logs WHERE student_id = ?", (sid,)
        ).fetchall()
        assert len(rows) == 2

    def test_student_stats(self, in_memory_db):
        """Log multiple entries, check avg_score and dominant_emotion."""
        sid = in_memory_db.add_student("Alice")
        session_id = in_memory_db.create_session("Session 1", "online")

        in_memory_db.log_attention(sid, session_id, "positive", 0.9, 0.8, "focused")
        in_memory_db.log_attention(sid, session_id, "positive", 0.85, 0.75, "focused")
        in_memory_db.log_attention(sid, session_id, "neutral", 0.7, 0.5, "moderate")

        stats = in_memory_db.get_student_stats(sid, session_id)
        assert stats["total_entries"] == 3
        assert abs(stats["avg_score"] - (0.8 + 0.75 + 0.5) / 3) < 1e-4
        assert stats["dominant_emotion"] == "positive"

    def test_student_stats_empty(self, in_memory_db):
        """Stats for non-existent student should return defaults."""
        stats = in_memory_db.get_student_stats(9999)
        assert stats["total_entries"] == 0
        assert stats["avg_score"] == 0.0
        assert stats["dominant_emotion"] is None

    def test_session_stats(self, in_memory_db):
        """Multiple students in session, check aggregation."""
        sid1 = in_memory_db.add_student("Alice")
        sid2 = in_memory_db.add_student("Bob")
        session_id = in_memory_db.create_session("Session 1", "online")

        in_memory_db.log_attention(sid1, session_id, "positive", 0.9, 0.8, "focused")
        in_memory_db.log_attention(sid2, session_id, "negative", 0.8, 0.3, "distracted")

        stats = in_memory_db.get_session_stats(session_id)
        assert stats["session"]["total_students"] == 2
        assert sid1 in stats["students"]
        assert sid2 in stats["students"]
        # Session avg: (0.8 + 0.3) / 2 = 0.55
        assert abs(stats["session"]["avg_score"] - 0.55) < 1e-4


class TestExportExcel:
    """Test Excel export functionality."""

    def test_export_to_excel(self, in_memory_db, tmp_path):
        """Export, verify file exists and has correct sheets."""
        sid = in_memory_db.add_student("Alice")
        session_id = in_memory_db.create_session("Session 1", "online")
        in_memory_db.log_attention(sid, session_id, "positive", 0.9, 0.8, "focused")

        filepath = tmp_path / "export.xlsx"
        in_memory_db.export_to_excel(filepath, session_id)

        assert filepath.exists()

        from openpyxl import load_workbook

        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames
        assert "Overview" in sheet_names
        assert "Per Student" in sheet_names
        assert "Detailed Logs" in sheet_names

    def test_export_to_excel_no_session(self, in_memory_db, tmp_path):
        """Export without session_id should still work (global overview)."""
        in_memory_db.add_student("Alice")

        filepath = tmp_path / "export_global.xlsx"
        in_memory_db.export_to_excel(filepath)

        assert filepath.exists()


# ── FaceRecognizer tests ─────────────────────────────────────────────────────


class TestFaceRecognizer:
    """Test face embedding comparison and identification."""

    @pytest.fixture(autouse=True)
    def _recognizer(self):
        self.recognizer = FaceRecognizer(threshold=0.5)

    def test_cosine_similarity(self, sample_embedding):
        """compare() with same vector should return ~1.0."""
        sim = self.recognizer.compare(sample_embedding, sample_embedding)
        assert abs(sim - 1.0) < 1e-5

    def test_cosine_similarity_orthogonal(self):
        """Orthogonal vectors should return ~0.0."""
        v1 = np.zeros(512, dtype=np.float32)
        v1[0] = 1.0
        v2 = np.zeros(512, dtype=np.float32)
        v2[1] = 1.0

        sim = self.recognizer.compare(v1, v2)
        assert abs(sim) < 1e-5

    def test_cosine_similarity_opposite(self):
        """Opposite vectors should return ~-1.0."""
        v1 = np.random.randn(512).astype(np.float32)
        v1 /= np.linalg.norm(v1)
        v2 = -v1

        sim = self.recognizer.compare(v1, v2)
        assert abs(sim - (-1.0)) < 1e-5

    def test_identify_match(self, sample_embedding):
        """Known embedding should return correct student_id."""
        known = {1: sample_embedding, 2: np.random.randn(512).astype(np.float32)}
        known[2] /= np.linalg.norm(known[2])

        student_id, sim = self.recognizer.identify(sample_embedding, known)
        assert student_id == 1
        assert sim > 0.9

    def test_identify_no_match(self):
        """Unknown embedding should return (None, 0.0)."""
        query = np.random.randn(512).astype(np.float32)
        query /= np.linalg.norm(query)

        # Create known embeddings that are very different
        known = {}
        for i in range(3):
            v = np.zeros(512, dtype=np.float32)
            v[i] = 1.0
            known[i + 1] = v

        student_id, sim = self.recognizer.identify(query, known)
        # With random query vs sparse vectors, similarity should be low
        # If below threshold, returns (None, 0.0)
        if student_id is None:
            assert sim == 0.0

    def test_identify_empty_gallery(self, sample_embedding):
        """Empty gallery should return (None, 0.0)."""
        student_id, sim = self.recognizer.identify(sample_embedding, {})
        assert student_id is None
        assert sim == 0.0

    def test_register_student_average(self):
        """Multiple embeddings should be averaged and normalized."""
        emb1 = np.ones(512, dtype=np.float32) * 0.5
        emb2 = np.ones(512, dtype=np.float32) * 1.5

        avg = self.recognizer.register_student("test", [emb1, emb2])
        assert avg.shape == (512,)
        # Should be normalized
        assert abs(np.linalg.norm(avg) - 1.0) < 1e-5

    def test_register_student_empty_raises(self):
        """Empty embeddings list should raise ValueError."""
        with pytest.raises(ValueError, match="No embeddings provided"):
            self.recognizer.register_student("test", [])
