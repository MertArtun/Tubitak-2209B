"""SQLite database for student records, face embeddings, sessions and attention logs."""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np

from configs.config import DB_PATH

logger = logging.getLogger(__name__)

# Embedding dimensionality used by ArcFace
_EMBEDDING_DIM = 512


class StudentDatabase:
    """Persistent storage for students, embeddings, sessions and attention logs."""

    def __init__(self, db_path: str | Path = DB_PATH) -> None:
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._conn: sqlite3.Connection | None = None
        self._connect()
        self._init_db()

    # ------------------------------------------------------------------
    # Connection helpers
    # ------------------------------------------------------------------

    def _connect(self) -> None:
        self._conn = sqlite3.connect(str(self.db_path))
        self._conn.row_factory = sqlite3.Row
        self._conn.execute("PRAGMA foreign_keys = ON")

    @property
    def conn(self) -> sqlite3.Connection:
        if self._conn is None:
            self._connect()
        return self._conn  # type: ignore[return-value]

    # ------------------------------------------------------------------
    # Schema
    # ------------------------------------------------------------------

    def _init_db(self) -> None:
        """Create tables if they do not exist."""
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS students (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT NOT NULL,
                email       TEXT,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS face_embeddings (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id  INTEGER NOT NULL REFERENCES students(id),
                embedding   BLOB NOT NULL,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT,
                mode        TEXT CHECK(mode IN ('online', 'face-to-face')),
                start_time  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                end_time    TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS attention_logs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id      INTEGER NOT NULL REFERENCES students(id),
                session_id      INTEGER NOT NULL REFERENCES sessions(id),
                timestamp       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                emotion         TEXT,
                confidence      REAL,
                attention_score REAL,
                attention_level TEXT
            );
            """
        )
        self.conn.commit()

    # ------------------------------------------------------------------
    # Students
    # ------------------------------------------------------------------

    def add_student(self, name: str, email: str | None = None) -> int:
        """Insert a new student and return its id."""
        cur = self.conn.execute(
            "INSERT INTO students (name, email) VALUES (?, ?)", (name, email)
        )
        self.conn.commit()
        return cur.lastrowid  # type: ignore[return-value]

    def get_student(self, student_id: int) -> dict | None:
        """Return student dict or None."""
        row = self.conn.execute(
            "SELECT * FROM students WHERE id = ?", (student_id,)
        ).fetchone()
        return dict(row) if row else None

    def list_students(self) -> list[dict]:
        """Return all students."""
        rows = self.conn.execute("SELECT * FROM students ORDER BY id").fetchall()
        return [dict(r) for r in rows]

    # ------------------------------------------------------------------
    # Face embeddings
    # ------------------------------------------------------------------

    def save_embedding(self, student_id: int, embedding: np.ndarray) -> None:
        """Store a face embedding as a binary blob."""
        blob = embedding.astype(np.float32).tobytes()
        self.conn.execute(
            "INSERT INTO face_embeddings (student_id, embedding) VALUES (?, ?)",
            (student_id, blob),
        )
        self.conn.commit()

    def get_embeddings(self, student_id: int) -> list[np.ndarray]:
        """Retrieve all embeddings for a given student."""
        rows = self.conn.execute(
            "SELECT embedding FROM face_embeddings WHERE student_id = ?",
            (student_id,),
        ).fetchall()
        return [
            np.frombuffer(row["embedding"], dtype=np.float32).copy()
            for row in rows
        ]

    def get_all_embeddings(self) -> dict[int, np.ndarray]:
        """Return student_id -> average (normalised) embedding for every student.

        Used as the gallery for face recognition.
        """
        rows = self.conn.execute(
            "SELECT student_id, embedding FROM face_embeddings"
        ).fetchall()

        buckets: dict[int, list[np.ndarray]] = {}
        for row in rows:
            sid = row["student_id"]
            emb = np.frombuffer(row["embedding"], dtype=np.float32).copy()
            buckets.setdefault(sid, []).append(emb)

        result: dict[int, np.ndarray] = {}
        for sid, embs in buckets.items():
            avg = np.mean(embs, axis=0).astype(np.float32)
            norm = np.linalg.norm(avg)
            if norm > 0:
                avg = avg / norm
            result[sid] = avg
        return result

    # ------------------------------------------------------------------
    # Sessions
    # ------------------------------------------------------------------

    def create_session(self, name: str, mode: str) -> int:
        """Create a new session and return its id."""
        cur = self.conn.execute(
            "INSERT INTO sessions (name, mode) VALUES (?, ?)", (name, mode)
        )
        self.conn.commit()
        return cur.lastrowid  # type: ignore[return-value]

    def end_session(self, session_id: int) -> None:
        """Set the end_time of a session to now."""
        self.conn.execute(
            "UPDATE sessions SET end_time = ? WHERE id = ?",
            (datetime.now().isoformat(), session_id),
        )
        self.conn.commit()

    # ------------------------------------------------------------------
    # Attention logs
    # ------------------------------------------------------------------

    def log_attention(
        self,
        student_id: int,
        session_id: int,
        emotion: str,
        confidence: float,
        attention_score: float,
        attention_level: str,
    ) -> None:
        """Insert an attention log entry."""
        self.conn.execute(
            """INSERT INTO attention_logs
               (student_id, session_id, emotion, confidence,
                attention_score, attention_level)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (student_id, session_id, emotion, confidence, attention_score, attention_level),
        )
        self.conn.commit()

    # ------------------------------------------------------------------
    # Statistics
    # ------------------------------------------------------------------

    def get_student_stats(
        self, student_id: int, session_id: int | None = None
    ) -> dict:
        """Compute aggregated attention stats for a student.

        Returns:
            Dict with avg_score, dominant_emotion, total_entries,
            and attention_distribution.
        """
        query = "SELECT * FROM attention_logs WHERE student_id = ?"
        params: list[Any] = [student_id]
        if session_id is not None:
            query += " AND session_id = ?"
            params.append(session_id)

        rows = self.conn.execute(query, params).fetchall()
        if not rows:
            return {
                "avg_score": 0.0,
                "dominant_emotion": None,
                "total_entries": 0,
                "attention_distribution": {},
            }

        scores = [r["attention_score"] for r in rows]
        emotions = [r["emotion"] for r in rows]
        levels = [r["attention_level"] for r in rows]

        # Emotion frequency
        emotion_counts: dict[str, int] = {}
        for e in emotions:
            emotion_counts[e] = emotion_counts.get(e, 0) + 1
        dominant_emotion = max(emotion_counts, key=emotion_counts.get)  # type: ignore[arg-type]

        # Attention level distribution
        level_dist: dict[str, int] = {}
        for lv in levels:
            level_dist[lv] = level_dist.get(lv, 0) + 1

        return {
            "avg_score": float(np.mean(scores)),
            "dominant_emotion": dominant_emotion,
            "total_entries": len(rows),
            "attention_distribution": level_dist,
        }

    def get_session_stats(self, session_id: int) -> dict:
        """Compute per-student and session-wide stats.

        Returns:
            Dict with 'students' (per-student stats) and 'session' averages.
        """
        rows = self.conn.execute(
            "SELECT DISTINCT student_id FROM attention_logs WHERE session_id = ?",
            (session_id,),
        ).fetchall()

        per_student: dict[int, dict] = {}
        all_scores: list[float] = []

        for row in rows:
            sid = row["student_id"]
            stats = self.get_student_stats(sid, session_id)
            per_student[sid] = stats
            if stats["total_entries"] > 0:
                all_scores.append(stats["avg_score"])

        session_avg = float(np.mean(all_scores)) if all_scores else 0.0

        return {
            "students": per_student,
            "session": {
                "avg_score": session_avg,
                "total_students": len(per_student),
            },
        }

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_to_excel(
        self, filepath: str | Path, session_id: int | None = None
    ) -> None:
        """Export attention data to an Excel file with multiple worksheets.

        Worksheets:
            - Overview: session-level summary.
            - Per Student: per-student aggregated stats.
            - Detailed Logs: raw attention log entries.
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Overview sheet ────────────────────────────────────────────
        ws_overview = wb.active
        ws_overview.title = "Overview"  # type: ignore[union-attr]

        if session_id is not None:
            session_row = self.conn.execute(
                "SELECT * FROM sessions WHERE id = ?", (session_id,)
            ).fetchone()
            stats = self.get_session_stats(session_id)
            overview_headers = ["Field", "Value"]
            overview_data = [
                ["Session ID", session_id],
                ["Session Name", session_row["name"] if session_row else ""],
                ["Mode", session_row["mode"] if session_row else ""],
                ["Start Time", session_row["start_time"] if session_row else ""],
                ["End Time", session_row["end_time"] if session_row else ""],
                ["Total Students", stats["session"]["total_students"]],
                ["Average Score", round(stats["session"]["avg_score"], 4)],
            ]
        else:
            overview_headers = ["Field", "Value"]
            total_students = self.conn.execute(
                "SELECT COUNT(*) AS cnt FROM students"
            ).fetchone()["cnt"]
            total_sessions = self.conn.execute(
                "SELECT COUNT(*) AS cnt FROM sessions"
            ).fetchone()["cnt"]
            total_logs = self.conn.execute(
                "SELECT COUNT(*) AS cnt FROM attention_logs"
            ).fetchone()["cnt"]
            overview_data = [
                ["Total Students", total_students],
                ["Total Sessions", total_sessions],
                ["Total Log Entries", total_logs],
            ]

        ws_overview.append(overview_headers)  # type: ignore[union-attr]
        for row in overview_data:
            ws_overview.append(row)  # type: ignore[union-attr]

        # ── Per Student sheet ─────────────────────────────────────────
        ws_students = wb.create_sheet("Per Student")
        student_headers = [
            "Student ID",
            "Name",
            "Avg Score",
            "Dominant Emotion",
            "Total Entries",
        ]
        ws_students.append(student_headers)

        students = self.list_students()
        for s in students:
            st = self.get_student_stats(s["id"], session_id)
            ws_students.append(
                [
                    s["id"],
                    s["name"],
                    round(st["avg_score"], 4),
                    st["dominant_emotion"] or "",
                    st["total_entries"],
                ]
            )

        # ── Detailed Logs sheet ───────────────────────────────────────
        ws_logs = wb.create_sheet("Detailed Logs")
        log_headers = [
            "Log ID",
            "Student ID",
            "Session ID",
            "Timestamp",
            "Emotion",
            "Confidence",
            "Attention Score",
            "Attention Level",
        ]
        ws_logs.append(log_headers)

        query = "SELECT * FROM attention_logs"
        params_list: list[Any] = []
        if session_id is not None:
            query += " WHERE session_id = ?"
            params_list.append(session_id)
        query += " ORDER BY timestamp"

        log_rows = self.conn.execute(query, params_list).fetchall()
        for lr in log_rows:
            ws_logs.append(
                [
                    lr["id"],
                    lr["student_id"],
                    lr["session_id"],
                    lr["timestamp"],
                    lr["emotion"],
                    lr["confidence"],
                    lr["attention_score"],
                    lr["attention_level"],
                ]
            )

        # Auto-width columns for all sheets
        for ws in [ws_overview, ws_students, ws_logs]:
            for col_idx in range(1, ws.max_column + 1):  # type: ignore[operator]
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(  # type: ignore[union-attr]
                    min_col=col_idx, max_col=col_idx
                ):
                    for cell in row:
                        val = str(cell.value) if cell.value is not None else ""
                        max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = max_len + 2  # type: ignore[index]

        wb.save(str(filepath))
        logger.info("Exported attention data to %s", filepath)

    # ------------------------------------------------------------------
    # Context manager & cleanup
    # ------------------------------------------------------------------

    def close(self) -> None:
        """Close the database connection."""
        if self._conn is not None:
            self._conn.close()
            self._conn = None

    def __enter__(self) -> "StudentDatabase":
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        self.close()
